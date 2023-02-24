from openpyxl import Workbook
import openpyxl
from openpyxl.styles import *
from openpyxl.worksheet.table import Table, TableStyleInfo
import pprint

text_file = open('employees.txt')
records = []
text_file.seek(0)
for record in text_file.readlines():
    records.append(record.rstrip("\n").split(";"))

workbook = Workbook()
filepath = "MyCompanyStaff.xlsx"
workbook.save(filepath)

# print(workbook.sheetnames)

#Rename worksheet
sheet = workbook['Sheet']
sheet.title = "Employees"
print(workbook.sheetnames)

sheet = workbook['Employees']

for row in records:
    sheet.append(row)

table = Table(displayName="Table", ref = "A1:G11")
print(openpyxl.worksheet.table.TABLESTYLES)

style = TableStyleInfo(name = "TableStyleMedium9", showRowStripes=True, showColumnStripes = True)

#apply styl to table
table.tableStyleInfo = style

sheet.add_table(table)
font = Font( bold=True, italic =True)

for cell_no in range(2,12):
    if int(sheet['G%s' % (cell_no)].value) > 55000:
        sheet['G%s' % (cell_no)].font = font

#Saving the changes made to the workbook
workbook.save(filepath)

#Closing the text file
text_file.close()

#Closing the workbook
workbook.close()
