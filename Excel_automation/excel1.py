import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
import pprint

workbook = openpyxl.load_workbook('AmazonReviews.xlsx')
sheet = workbook.active
print(workbook.sheetnames)
# print(sheet.title)
cell = sheet["C2"]
# print(cell.value)
"""
ALTERNATIVELY
"""
cell1 = sheet.cell(row=2,column=3)
# print(cell1.value)
# for i in range(1,6,2):
#     print(sheet.cell(row = i, column=4).value)

cell_range = sheet["A":"C"]
# print(cell_range)
# workbook = sheet('AmazonReviews')
# for file in workbook:
#     print(file)


# for row in sheet.iter_rows(min_row=1, max_row=10,min_col=1,max_col=10,values_only=True):
#     print(row)
# for col in sheet.iter_cols(min_row=1, max_row=10,min_col=1,max_col=10,values_only=True):
#     print(col)

# for cols in sheet.columns:
#     for col in cols:
#         print(col.value)

#to call second column
# for col in list(sheet.columns)[1]:
#     print(col.value)

# for cells in sheet["A1":"C10"]:
#     for cell in cells:
#         print(cell.coordinate, cell.value)

# # print(get_column_letter(sheet.max_column))
# print(get_column_letter(sheet.min_column))
# print(column_index_from_string("J"))
for value in sheet.iter_rows(min_row=1, max_row=1,values_only=True):
    print(value)
reviews = {}
for row in sheet.iter_rows(min_row=2,min_col=1, max_col=8, values_only=True):
    print(row)
    review_id = row[0]
    review = {
        'profileName': row[1],
        "title": row[4],
        "rating": row[5]
    }
    reviews[review_id] = review
pprint.pprint(reviews)
